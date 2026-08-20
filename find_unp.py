#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Автоматический поиск УНП/регномера организаций и ИП (Беларусь) по названию.

Источник данных: открытый API Единого государственного регистра юридических
лиц и индивидуальных предпринимателей (ЕГР) — https://www.egr.gov.by/

Принцип работы
--------------
1. Читает список названий из xlsx/csv (столбец задаётся через --column).
2. Для каждого названия запрашивает ЕГР поиск по наименованию.
3. Нормализует названия (Ё→Е, регистр, лишние пробелы, кавычки, ООО/ЗАО/ИП).
4. Сопоставляет найденные варианты с исходным названием с помощью нечёткого
   сравнения (rapidfuzz) и выбирает лучший + считает балл уверенности.
5. Сохраняет промежуточные результаты в кэш (JSON) — повторный запуск
   продолжается с того же места, не опрашивая ЕГР повторно.
6. Выгружает итоговый Excel: исходное название, УНП, найденное название,
   статус в реестре, балл, решение (auto / review / not_found), кандидаты.

ВАЖНО: сервер egr.gov.by блокирует некоторые зарубежные/"датацентр" IP-адреса.
Если поиск не идёт, запускайте скрипт с компьютера, у которого есть доступ к
egr.gov.by (например, из белорусской сети). Используйте ключ --probe, чтобы
посмотреть "сырой" ответ сервера и при необходимости подкорректировать парсер.

Изменения для дополнительного поиска при получении нулевого результата первоначального запроса
# 1. Сначала выполняется обычный поиск egr_search(search_query).
# 2. Если он вернул 0 кандидатов и это ЮЛ, выполняется
#    дополнительный egr_search_by_words(search_query, cache).
# 3. Дополнительный поиск НЕ меняет поисковый запрос.
# 4. Все найденные кандидаты снова проходят через тот же best_match().
# 5. Если дополнительный поиск также ничего не дал -> not_found.
# 6. Если найден кандидат -> решение определяется обычным decide(score).
# 7. В поле "Источник" фиксируется, каким способом получен результат.

Установка зависимостей:
    pip install requests openpyxl rapidfuzz

Запуск:
    python find_unp.py input.xlsx --column name --output result.xlsx
    python find_unp.py input.csv  --column 1            --output result.xlsx
    python find_unp.py input.xlsx --column name --probe     # отладка одного запроса


API взят из New_egr_API.docx:
    http://egr.gov.by/api/v2/egr/getShortInfoByRegName/{name}

ВАЖНО:
В предоставленном описании API поле называется NGRN — "Регистрационный номер".
Документ не использует термин "УНП". Поэтому результат помечается как
"УНП (NGRN)", чтобы не выдавать предположение за название поля API.
"""

import argparse
import csv
import json
import re
import sys
import time
from pathlib import Path
from urllib.parse import quote

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from rapidfuzz import fuzz


EGR_BASE_URL = "http://egr.gov.by/api/v2/egr"
EGR_SEARCH_URL = EGR_BASE_URL + "/getShortInfoByRegName/{name}"

EGR_TIMEOUT = 30
EGR_DELAY = 0.5
EGR_RETRIES = 3

SCORE_ACCEPT = 90
SCORE_REVIEW = 75
SCORE_MARGIN = 10

LEGAL_FORMS = [
    "ооо",
    "чтуп",
    "оао",
    "ип",
    "чуп",
    "одо",
    "чп",
    "чтпуп",
    "зао",
    "чпуп",
    "иооо",
    "уп",
    "руп",
    "пк",
    "пуп",
    "чтсуп",
    "киуп",
    "куп",
    "птчуп",
    "сба засо",
    "мо оао",
    "тп",
    "туп",
    "уз",
    "уо",
    "утк",
    "учтпп",
    "учтп",
    "чптуп",
    "чсуп",
    "чткп",
    "сооо",
    "мукп",
    "филиал зао",
    "филиал оао",
    "филиал",
    "частное предприятие",
    "общество с ограниченной ответственностью",
    "закрытое акционерное общество",
    "открытое акционерное общество",
    "частное унитарное предприятие",
    "коммунальное унитарное предприятие",
]
"""
QUOTE_RE = Quote Regular Expression (Регулярное выражение для кавычек)
-Эта строка находит все виды кавычек в тексте
WS_RE = Whitespace Regular Expression (Регулярное выражение для пробельных символов)
-Эта строка находит любые последовательности пробельных символов
"""
QUOTE_RE = re.compile(r'[«»“”\'"`]')
WS_RE = re.compile(r"\s+")

def debug_ip_candidates(input_name, candidates):
    """
    Диагностика сопоставления ИП по фамилии и инициалам.

    Ничего не выбирает и не изменяет.
    Показывает, как исходное ФИО и каждый кандидат
    разбираются на фамилию и инициалы.
    """

    def fio_parts(value):
        if not value:
            return "", ""

        s = str(value).strip().lower().replace("ё", "е")
        s = re.sub(r"\s+", " ", s)

        words = s.replace(",", " ").split()

        if not words:
            return "", ""

        surname = words[0]
        initials = ""

        for word in words[1:]:
            # Слово состоит из инициалов:
            # "т.в.", "т. в.", "т.в"
            if "." in word:
                letters = re.findall(r"[а-яa-z]", word)
                initials += "".join(letters)
                continue

            # Полное имя/отчество:
            # "татьяна" -> т
            # "викторовна" -> в
            letters = re.findall(r"[а-яa-z]", word)

            if letters:
                initials += letters[0]

        return surname, initials

    input_surname, input_initials = fio_parts(input_name)

    print("=" * 80)
    print("ДИАГНОСТИКА КАНДИДАТОВ ИП")
    print("=" * 80)

    print(f"Исходное ФИО : {input_name}")
    print(f"Фамилия      : {input_surname}")
    print(f"Инициалы     : {input_initials}")
    print(f"Кандидатов   : {len(candidates)}")
    print()

    matches = []

    for i, candidate in enumerate(candidates, start=1):

        vfio = candidate.get("vfio", "") or ""

        # ЮЛ среди результатов ИП нам неинтересны.
        if not vfio:
            continue

        surname, initials = fio_parts(vfio)

        surname_match = surname == input_surname
        initials_match = initials == input_initials

        if surname_match and initials_match:
            match = "ФАМИЛИЯ + ИНИЦИАЛЫ"
            matches.append(candidate)

        elif surname_match:
            match = "ТОЛЬКО ФАМИЛИЯ"

        else:
            match = ""

        # Показываем только кандидатов с совпадающей фамилией. Остальные не представляют интереса.
        if surname_match:
            print(
                f"{len(matches):>2}. "
                f"УНП={candidate.get('unp', '')} | "
                f"ФИО={vfio}"
            )
            print(
                f"    Фамилия: {surname} | "
                f"Инициалы: {initials} | "
                f"{match} | "
                f"Статус: {candidate.get('status', '')}"
            )

    print()
    print("-" * 80)
    print(f"Совпадений фамилии + инициалов: {len(matches)}")
    print("-" * 80)

    if matches:
        for i, candidate in enumerate(matches, start=1):
            print(
                f"{i}. "
                f"УНП={candidate.get('unp', '')} | "
                f"ФИО={candidate.get('vfio', '')} | "
                f"Статус={candidate.get('status', '')}"
            )
    else:
        print("Точных совпадений по фамилии + инициалам нет.")

    print("=" * 80)

    return matches
def normalize(name, tip_org):
    """Нормализация поискового запроса."""
    # print("->normalize(): ",name)
    if not name:
        return ""

    s = str(name).lower().strip().replace("ё", "е")
    s = QUOTE_RE.sub(" ", s)

    # Отбрасываем правовые формы.
    for lf in LEGAL_FORMS:
        s = re.sub(rf"\b{re.escape(lf)}\b", " ", s)

    s = re.sub(r"[^\w\s.№-]", " ", s, flags=re.UNICODE)
    s = WS_RE.sub(" ", s).strip()

    # Для ИП: первое слово — фамилия; инициалы после фамилии отбрасываем.
    if str(tip_org).strip().upper() == "ИП":
        words = s.split()
        if words:
            result = [words[0]]
            for word in words[1:]:
                # Инициал: одна буква, возможно с точкой.
                if len(word.replace(".", "")) <= 1:
                    continue
                # Сочетание инициалов также отбрасываем.
                if all(len(part) == 1 for part in word.split(".") if part):
                    continue
                result.append(word)
            s = " ".join(result)
    # print("normalize()->: ",s.strip())
    return s.strip()

def _as_dict_list(data):
    """Привести разные варианты JSON-ответа к списку словарей."""
    if data is None:
        return []
    if isinstance(data, dict):
        # Частые варианты: {"content": [...]}, {"data": [...]}, {"result": [...]}
        for key in ("content", "data", "result", "items", "rows", "list"):
            value = data.get(key)
            if isinstance(value, list):
                return [x for x in value if isinstance(x, dict)]
            if isinstance(value, dict):
                nested = _as_dict_list(value)
                if nested:
                    return nested
        return [data]
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    return []


def parse_egr_response(data):
    """
    Парсер именно схемы из New_egr_API.docx:
      ngrn   — регистрационный номер
      nsi00219 — состояние
      vfio   — ФИО ИП
      vnaim  — полное наименование ЮЛ
      vn     — сокращенное наименование ЮЛ
      vfn    — фирменное наименование ЮЛ
    """
    candidates = []
    for item in _as_dict_list(data):
        ngrn = item.get("ngrn", item.get("NGRN"))
        if ngrn is None:
            continue

        # API описывает ngrn как integer; сохраняем ведущие нули, если сервер
        # неожиданно вернул его строкой.
        ngrn = str(ngrn).strip()
        if not re.fullmatch(r"\d{1,12}", ngrn):
            continue

        status = item.get("nsi00219", item.get("NSI00219"))
        if isinstance(status, dict):
            status = (
                status.get("vnsostk")
                or status.get("VNSOSTK")
                or status.get("nksost")
                or status.get("NKSOST")
                or str(status)
            )

        vfio = item.get("vfio", item.get("VFIO"))
        vn = item.get("vn", item.get("VN"))
        vnaim = item.get("vnaim", item.get("VNAIM"))
        vfn = item.get("vfn", item.get("VFN"))

        # Сохраняем отдельные поля API.
        # Для ИП при сопоставлении и выводе используем VFIO,
        # для ЮЛ — VN (сокращенное наименование).
        if vfio or vn or vnaim or vfn:
            candidates.append({
                "unp": ngrn,
                "vfio": "" if vfio is None else str(vfio).strip(),
                "vn": "" if vn is None else str(vn).strip(),
                "vnaim": "" if vnaim is None else str(vnaim).strip(),
                "vfn": "" if vfn is None else str(vfn).strip(),
                "status": "" if status is None else str(status).strip(),
                "raw": item,
            })
    return candidates


_session = requests.Session()
_session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/151.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
})


def egr_search(name, debug=False):
    # Для path-параметра используем quote, а не quote_plus:
    # пробел должен передаваться как %20, а не как '+'.
    # debug=True — это инструмент разработчика
    # Статусы 204 (No Content) и 404 (Not Found)
    encoded = quote(str(name), safe="")
    url = EGR_SEARCH_URL.format(name=encoded)
    last_error = None

    for attempt in range(1, EGR_RETRIES + 1):
        try:
            response = _session.get(url, timeout=EGR_TIMEOUT)

            if debug:
                print("URL:", response.url)
                print("HTTP:", response.status_code)
                print("RAW:", response.text[:3000])

            if response.status_code == 204:
                return []

            if response.status_code == 404:
                return []

            if response.status_code == 200:
                try:
                    data = response.json()
                except ValueError:
                    raise RuntimeError(
                        "HTTP 200, но тело ответа не является JSON: "
                        + response.text[:500]
                    )
                candidates = parse_egr_response(data)
                if debug:
                    print("CANDIDATES:", json.dumps(
                        candidates, ensure_ascii=False, indent=2
                    )[:5000])
                return candidates

            last_error = f"HTTP {response.status_code}: {response.text[:500]}"

        except requests.RequestException as exc:
            last_error = str(exc)

        if attempt < EGR_RETRIES:
            time.sleep(attempt)

    raise RuntimeError(last_error or "Неизвестная ошибка запроса ЕГР")

def egr_search_by_words(search_query, cache):
    """
    Дополнительный поиск кандидатов ЮЛ по отдельным словам
    поискового запроса.

    Функция предназначена только для случая, когда обычный
    egr_search(search_query) вернул пустой список.

    Функция НЕ:
      - рассчитывает score;
      - вызывает best_match();
      - принимает решение auto/review/low/manual_multiple;
      - изменяет исходный поисковый запрос.

    Она только:
      1. разбирает search_query на слова;
      2. выполняет поиск по каждому слову;
      3. объединяет найденных кандидатов;
      4. удаляет дубли по УНП;
      5. сохраняет результаты отдельных запросов в cache.

    Параметры:
      search_query -- исходный поисковый запрос;
      cache        -- словарь существующего кэша.

    Возвращает:
      объединённый список кандидатов.
    """
    # print("Исходный запрос egr_search_by_words:", type(search_query), search_query)
    search_query = str(search_query or "").strip()

    # Пустой запрос
    if not search_query:
        return []


    # 1. РАБОТАЕМ СО СТРОКОЙ: Разделяем точку, если после нее идет буква (г.Могилеве -> г. Могилеве)
    prepared_query = re.sub(r'\.(?=[^\s])', '. ', search_query)

    # 2. РАБОТАЕМ СО СТРОКОЙ: Сохраняем буквы, цифры, пробелы, точки, а также знаки: №, / и -
    clean_query = re.sub(r'[^\w\s.№/\-]', ' ', prepared_query)

    # 3. ТОЛЬКО ТЕПЕРЬ разбиваем очищенную строку по пробелам
    raw_words = clean_query.split()

    # 4. Применяем правила фильтрации:
    # - длина больше 1 символа
    # - не заканчивается на точку
    words = [w for w in raw_words if len(w) > 1 and not w.endswith('.')]
    # print("Результат фильтрации egr_search_by_words:", words)
    
    # Если слово только одно, дополнительный поиск
    # не имеет смысла: исходный поиск уже был выполнен.
    if len(words) <= 1:
        return []

    # Словарь:
    #   ключ   = УНП
    #   значение = полный объект кандидата,
    #               возвращённый egr_search()
    candidates_by_unp = {}

    for word in words:

        # Кэш отдельного поискового запроса.
        #
        # Используем отдельный ключ, чтобы запрос
        # "ЮЛ|торцовых" не смешивался с исходным
        # "ЮЛ|нпп завод торцовых уплотнений".
        cache_key = f"ЮЛ|{word}"

        if cache_key in cache:
            candidates = cache[cache_key]
        else:
            try:
                candidates = egr_search(word)
                cache[cache_key] = candidates
            except TypeError as e:
                print("\n!!! КРИТИЧЕСКАЯ ОШИБКА ВНУТРИ egr_search !!!")
                print(f"Сбой произошло на слове: '{word}' (тип: {type(word)})")
                raise e

        # Если по отдельному слову ничего не найдено,
        # просто переходим к следующему слову.
        if not candidates:
            continue

        # Добавляем кандидатов в общий список.
        # УНП уникален, поэтому используем его
        # для удаления дублей.
        for candidate in candidates:

            unp = str(candidate.get("unp", "") or "").strip()

            # Кандидат без УНП не имеет смысла для
            # последующего выбора.
            if not unp:
                continue

            if unp not in candidates_by_unp:
                candidates_by_unp[unp] = candidate

    return list(candidates_by_unp.values())

def candidate_name(candidate, tip_org):
    """Название кандидата, используемое для сопоставления и вывода."""
    tip_org = str(tip_org or "").strip().upper()

    if tip_org == "ИП":
        return candidate.get("vfio", "")
    if tip_org == "ЮЛ":
        return candidate.get("vn", "")

    return ""


def best_match(candidates, query, tip_org):
    """
    1.Выбор кандидата.

    ИП:
      - только VFIO;
      - перед разбором ФИО удаляется правовая форма "ИП";
      - строгое совпадение фамилии и инициалов;
      - 0 -> not_found;
      - 1 -> auto;
      - >1 -> manual_multiple.

    ЮЛ:
      - существующий fuzzy-механизм по VN.
    
    2.Выбирает лучшего кандидата из нескольких по score.
    Возвращает:
        best       - лучший кандидат либо None
        second     - ближайший конкурент либо None
        best_score - score лучшего кандидата
        margin     - разница между первым и вторым кандидатом

    Для одного кандидата margin = None.
    """
    tip_org = str(tip_org or "").strip().upper()

    if tip_org == "ИП":

        def fio_parts(value):
            if not value:
                return "", ""

            s = str(value).strip().lower().replace("ё", "е")
            s = re.sub(r"\s+", " ", s)

            # Для входного значения "ИП Вишнякова Т.В."
            # удаляем правовую форму ИП перед разбором ФИО.
            words = s.replace(",", " ").split()
            if words and words[0] == "ип":
                words = words[1:]

            if not words:
                return "", ""

            surname = words[0]
            initials = ""

            for word in words[1:]:
                # Готовые инициалы: Т.В., Т. В., Т.В
                if "." in word:
                    letters = re.findall(r"[а-яa-z]", word)
                    initials += "".join(letters)
                    continue

                # Полное имя/отчество: Татьяна -> Т
                letters = re.findall(r"[а-яa-z]", word)
                if letters:
                    initials += letters[0]

            return surname, initials

        input_surname, input_initials = fio_parts(input_name)
        matched = []

        for candidate in candidates:
            vfio = candidate.get("vfio", "") or ""

            if not vfio:
                continue

            candidate_surname, candidate_initials = fio_parts(vfio)

            if (
                candidate_surname == input_surname
                and candidate_initials == input_initials
            ):
                matched.append({
                    "unp": candidate.get("unp", ""),
                    "name": vfio,
                    "status": candidate.get("status", ""),
                    "score": 100.0,
                })

        if not matched:
            return None, 0.0, []

        if len(matched) == 1:
            return matched[0], 100.0, matched

        # Несколько кандидатов: ничего автоматически не выбираем.
        return None, 100.0, matched

    # ЮЛ: существующий fuzzy-механизм по VN
    norm_in = normalize(input_name, tip_org)
    ranked = []

    for candidate in candidates:
        match_name = candidate_name(candidate, tip_org)

        if not match_name:
            continue

        score = fuzz.token_sort_ratio(
            norm_in,
            normalize(match_name, tip_org)
        )

        ranked.append({
            "unp": candidate.get("unp", ""),
            "name": match_name,
            "status": candidate.get("status", ""),
            "score": round(score, 1),
        })

    ranked.sort(key=lambda x: x[0], reverse=True)

    if not ranked:
        return None, None, 0.0, None

    best_score, best = ranked[0]

    if len(ranked) > 1:
        second_score, second = ranked[1]
        margin = best_score - second_score
    else:
        second = None
        margin = None

    return best, second, best_score, margin

def decide(score):
    if score >= SCORE_ACCEPT:
        return "auto"
    if score >= SCORE_REVIEW:
        return "review"
    return "low"


def _find_column(headers, column):
    """Найти индекс столбца по имени или 1-based номеру."""
    if isinstance(column, str) and not column.isdigit():
        matches = [
            i for i, h in enumerate(headers)
            if h.lower() == column.strip().lower()
        ]
        if not matches:
            raise ValueError(
                f"Столбец {column!r} не найден. Заголовки: {headers}"
            )
        return matches[0]
    return int(column) - 1


def read_names(
    path,
    id_column="id_r_raspost",
    type_column="TIP_ORG",
    name_column="ORG_NAME",
):
    """
    Прочитать входной XLSX.

    Возвращает ВСЕ строки с данными, включая строки с пустым ORG_NAME.
    Для каждой строки возвращается:
        (номер_строки_excel, ID, TIP_ORG, ORG_NAME)

    Номер строки Excel сохраняется только для диагностики.
    Основным идентификатором записи является id_r_raspost.
    """
    p = Path(path)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError("Для этой версии ожидается XLSX/XLSM")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []

    headers = ["" if x is None else str(x).strip() for x in header]

    id_idx = _find_column(headers, id_column)
    type_idx = _find_column(headers, type_column)
    name_idx = _find_column(headers, name_column)

    result = []

    for row_no, row in enumerate(rows, start=2):
        max_idx = max(id_idx, type_idx, name_idx)
        if max_idx >= len(row):
            # Не теряем строку из-за отсутствующего хвостового значения.
            # Но без ID невозможно однозначно идентифицировать запись.
            raise ValueError(
                f"Строка Excel {row_no} короче ожидаемой структуры."
            )

        record_id = row[id_idx]
        tip_org = row[type_idx]
        name = row[name_idx]

        record_id = "" if record_id is None else str(record_id).strip()
        tip_org = "" if tip_org is None else str(tip_org).strip()
        name = "" if name is None else str(name).strip()

        result.append((row_no, record_id, tip_org, name))

    wb.close()
    return result

def load_cache(path):
    p = Path(path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_cache(path, cache):
    p = Path(path)
    tmp = Path(str(p) + ".tmp")
    tmp.write_text(
        json.dumps(cache, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    tmp.replace(p)


def _cache_key(tip_org, name):
    """Ключ кэша учитывает тип субъекта и исходное название."""
    return f"{tip_org}|{name}"


def make_rows(records, cache_path="unp_cache.json"):
    """
    Основная обработка records.

    Правила:
      ФЛ -> inn_search(), без запроса УНП.
      Пустой ORG_NAME -> empty_name, строка сохраняется.
      ЮЛ/ИП -> обычный поиск ЕГР.
      Если обычный поиск ЮЛ дал 0 кандидатов ->
          дополнительный поиск по отдельным словам
          через egr_search_by_words().
      Ликвидированные организации не исключаются.

      Выбор кандидата:
        - один кандидат:
            score >= SCORE_ACCEPT -> auto
            score >= SCORE_REVIEW -> review
            иначе -> low
        - несколько кандидатов:
            score >= SCORE_ACCEPT AND margin >= SCORE_MARGIN -> auto
            score >= SCORE_REVIEW -> review
            иначе -> low

      Найденное название: ИП=VFIO, ЮЛ=VN.
    """
    cache = load_cache(cache_path)
    rows = []

    for rec in records:
        row_number, record_id, tip_org, name = rec

        tip_org = str(tip_org or "").strip().upper()
        name = str(name or "").strip()

        # ---------------------------------------------------------
        # Пустое наименование
        # ---------------------------------------------------------
        if not name:
            rows.append({
                "ID": record_id,
                "TIP_ORG": tip_org,
                "Исходное название": name,
                "Поисковый запрос": "",
                "УНП": "",
                "Найденное название": "",
                "Статус": "",
                "Балл": 0.0,
                "Решение": "empty_name",
                "Источник": "",
            })
            continue

        # ---------------------------------------------------------
        # Физическое лицо: УНП не ищем
        # ---------------------------------------------------------
        if tip_org == "ФЛ":
            inn_search(name)

            rows.append({
                "ID": record_id,
                "TIP_ORG": tip_org,
                "Исходное название": name,
                "Поисковый запрос": "",
                "УНП": "",
                "Найденное название": "",
                "Статус": "",
                "Балл": 0.0,
                "Решение": "inn_search",
                "Источник": "",
            })
            continue

        # ---------------------------------------------------------
        # ЮЛ / ИП
        # ---------------------------------------------------------
        search_query = normalize(name, tip_org)

        if not search_query:
            rows.append({
                "ID": record_id,
                "TIP_ORG": tip_org,
                "Исходное название": name,
                "Поисковый запрос": search_query,
                "УНП": "",
                "Найденное название": "",
                "Статус": "",
                "Балл": 0.0,
                "Решение": "not_found",
                "Источник": "",
            })
            continue

        # ---------------------------------------------------------
        # Первоначальный поиск / кэш
        # ---------------------------------------------------------
        cache_key = f"{tip_org}|{search_query}"

        if cache_key in cache:
            candidates = cache[cache_key]
            source = "кэш"
        else:
            candidates = egr_search(search_query)
            cache[cache_key] = candidates
            source = "запрос ЕГР"

        # ---------------------------------------------------------
        # ДОПОЛНИТЕЛЬНЫЙ ПОИСК
        #
        # Выполняется только если:
        #   - тип организации ЮЛ;
        #   - первоначальный поиск дал 0 кандидатов.
        # ---------------------------------------------------------
        if tip_org == "ЮЛ" and not candidates:
            word_candidates = egr_search_by_words(
                search_query,
                cache
            )

            if word_candidates:
                candidates = word_candidates
                source = "дополнительный поиск по словам"
            else:
                source = "поиск ЕГР + поиск по словам"

        # ---------------------------------------------------------
        # Выбор кандидата
        #
        # best_match() должен возвращать:
        #   best, second, score, margin
        #
        # Для одного кандидата:
        #   second = None
        #   margin = None
        #
        # Для нескольких кандидатов:
        #   margin = score(best) - score(second)
        # ---------------------------------------------------------
        best, second, score, margin = best_match(
            candidates,
            search_query,
            tip_org
        )

        # ---------------------------------------------------------
        # Кандидат найден
        # ---------------------------------------------------------
        if best is not None:
            unp = best.get("unp", "")
            found_name = best.get("name", "")
            status = best.get("status", "")
            result_score = best.get("score", score)

            # Один кандидат: достаточно самого score.
            if second is None:
                decision = decide(score)

            # Несколько кандидатов:
            # auto только при высоком score И достаточном отрыве
            # от ближайшего конкурента.
            else:
                if (
                    score >= SCORE_ACCEPT
                    and margin >= SCORE_MARGIN
                ):
                    decision = "auto"
                elif score >= SCORE_REVIEW:
                    decision = "review"
                else:
                    decision = "low"

        # ---------------------------------------------------------
        # Кандидат не выбран
        # ---------------------------------------------------------
        else:
            decision = "not_found"
            unp = ""
            found_name = ""
            status = ""
            result_score = 0.0

        # ---------------------------------------------------------
        # Итоговая строка
        # ---------------------------------------------------------
        rows.append({
            "ID": record_id,
            "TIP_ORG": tip_org,
            "Исходное название": name,
            "Поисковый запрос": search_query,
            "УНП": unp,
            "Найденное название": found_name,
            "Статус": status,
            "Балл": round(result_score, 1),
            "Решение": decision,
            "Источник": source,
        })

    save_cache(cache_path, cache)
    return rows

def test_single_record(records, row_number, cache_path="unp_cache.json"):
    """
    Тестовый режим для одной строки Excel.

    row_number — фактический номер строки Excel, включая строку заголовка.
    Запрос выполняется только для выбранной строки.
    Результат возвращается в том же формате 9 колонок, что и result.xlsx.
    """
    selected = [r for r in records if r[0] == row_number]

    if not selected:
        raise ValueError(
            f"Строка Excel {row_number} не найдена."
        )

    row = selected[0]
    row_no, record_id, tip_org, name = row

    print("\n=== ТЕСТ ОДНОЙ СТРОКИ ===")
    print("Строка Excel:", row_no)
    print("ID:", record_id)
    print("TIP_ORG:", tip_org)
    print("ORG_NAME:", name)

    if str(tip_org).strip().upper() not in ("ЮЛ", "ИП"):
        raise ValueError(
            "Тестовый режим для одной строки предназначен для ЮЛ или ИП."
        )

    if not name or name == "#нет#":
        result = {
            "unp": "",
            "matched": "",
            "status": "",
            "score": 0.0,
            "decision": "empty_name",
            "candidates": [],
            "error": "",
        }
        search_query = ""
    else:
        search_query = normalize(name, tip_org)
        print("Нормализованный поисковый запрос:", search_query)
        cache = load_cache(cache_path)
        key = _cache_key(tip_org, name)

        # В тестовом режиме кэш используем, но явно показываем это.
        if key in cache:
            print("Источник результата: кэш")
            result = cache[key]
        else:
            print("Источник результата: запрос ЕГР")
            candidates = egr_search(search_query, debug=True)
            best, score, ranked = best_match(
                name,
                candidates,
                tip_org,
            )

            if best is not None:
                decision = decide(score)
                unp = best.get("unp", "")
                matched = best.get("name", "")
                status = best.get("status", "")
            elif ranked:
                decision = "manual_multiple"
                unp = ""
                matched = ""
                status = ""
            else:
                decision = "not_found"
                unp = ""
                matched = ""
                status = ""

            result = {
                "unp": unp,
                "matched": matched,
                "status": status,
                "score": score,
                "decision": decision,
                "candidates": ranked,
                "error": "",
            }

            cache[key] = result
            save_cache(cache_path, cache)

    result_row = [
        record_id,
        tip_org,
        name,
        search_query,
        result.get("unp", ""),
        result.get("matched", ""),
        result.get("status", ""),
        result.get("score", 0),
        result.get("decision", ""),
    ]

    print("\n=== РЕЗУЛЬТАТ ===")
    print("УНП:", result_row[4] or "—")
    print("Найденное название:", result_row[5] or "—")
    print("Статус:", result_row[6] or "—")
    print("Балл:", result_row[7])
    print("Решение:", result_row[8])

    print("\n=== КАНДИДАТЫ ===")
    if result.get("candidates"):
        for n, candidate in enumerate(result["candidates"], start=1):
            print(
                f"{n}. УНП={candidate.get('unp','')} | "
                f"Название={candidate.get('name','')} | "
                f"Балл={candidate.get('score',0)} | "
                f"Статус={candidate.get('status','')}"
            )
    else:
        print("Кандидатов нет.")

    return [result_row]

def write_excel(rows, output):
    """Записать рабочий результат поиска УНП."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "УНП"

    headers = [
        "ID",
        "Тип",
        "Исходное название",
        "Поисковый запрос",
        "УНП",
        "Найденное название",
        "Статус",
        "Балл",
        "Решение",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(color="FFFFFF", bold=True)

    fills = {
        "auto": PatternFill("solid", fgColor="C6EFCE"),
        "review": PatternFill("solid", fgColor="FFEB9C"),
        "low": PatternFill("solid", fgColor="FFC7CE"),
        "not_found": PatternFill("solid", fgColor="FFC7CE"),
        "error": PatternFill("solid", fgColor="FFC7CE"),
        "empty_name": PatternFill("solid", fgColor="D9E1F2"),
    }

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
    for row in rows:
        ws.append([
            row.get("ID", ""),
            row.get("TIP_ORG", ""),
            row.get("Исходное название", ""),
            row.get("Поисковый запрос", ""),
            row.get("УНП", ""),
            row.get("Найденное название", ""),
            row.get("Статус", ""),
            row.get("Балл", ""),
            row.get("Решение", ""),
        ])
        # ws.append(row) - openpyxl при ws.append(dict) воспринимает ключи словаря как имена Excel-колонок.
        fill = fills.get(row.get("Решение", ""))
        # fill = fills.get(row[8]) взять элемент с индексом 8 - взять значение по ключу "Решение".
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    widths = [18, 10, 55, 55, 15, 55, 20, 10, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Сводка сохраняет информацию о том, что произошло при обработке.
    summary = wb.create_sheet("Сводка")
    from collections import Counter

    # counter = Counter(row[8] for row in rows)
    # counter = Counter(row["Решение"] for row in rows)
    
    counter = Counter()

    for row in rows:
        decision = row["Решение"]

        if decision.startswith("manual_multiple"):
            decision = "manual_multiple"

        counter[decision] += 1
    
    summary.append(["Решение", "Количество", "Доля"])
    total = len(rows)

    for key in ("auto", "review", "low", "manual_multiple", "not_found", "empty_name", "error"):
        if counter[key]:
            summary.append([
                key,
                counter[key],
                f"{counter[key] / total:.1%}" if total else "0%",
            ])

    summary.append(["ВСЕГО", total, "100%"])

    wb.save(output)
    

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="input.xlsx")
    parser.add_argument("--id-column", default="id_r_raspost")
    parser.add_argument("--type-column", default="TIP_ORG")
    parser.add_argument("--name-column", default="ORG_NAME")
    parser.add_argument("--output", default="result.xlsx")
    parser.add_argument("--cache", default="unp_cache.json")
    parser.add_argument(
        "--probe",
        action="store_true",
        help="Показать сырой ответ API для первой непустой строки.",
    )
    parser.add_argument(
        "--test-row",
        type=int,
        default=None,
        help="Тестировать одну конкретную строку Excel (номер строки, включая заголовок).",
    )
    parser.add_argument(
        "--test-output",
        default="result_test.xlsx",
        help="Файл результата тестового режима.",
    )
    args = parser.parse_args()

    records = read_names(
        args.input,
        id_column=args.id_column,
        type_column=args.type_column,
        name_column=args.name_column,
    )
    print(f"Будет обработано строк: {len(records)}")

    if args.probe:
        if not records:
            raise SystemExit("Нет строк для теста.")

        # Ищем первую строку с непустым названием.
        probe_record = next(
            (record for record in records if record[3] and record[3] != "#нет#"),
            None,
        )
        if probe_record is None:
            raise SystemExit("Нет непустых ORG_NAME для теста.")

        print("ID:", probe_record[1])
        print("Тип:", probe_record[2])
        print("Тест:", probe_record[3])
        egr_search(probe_record[3], debug=True)
        return

    if args.test_row is not None:
        rows = test_single_record(
            records,
            args.test_row,
            args.cache,
        )
        write_excel(rows, args.test_output)
        print("\nТестовый result.xlsx не перезаписывается.")
        print("Результат теста:", args.test_output)
        return

    rows = make_rows(records, args.cache)
    write_excel(rows, args.output)
    
    found = sum(row["Решение"] in ("auto", "review") for row in rows)
    empty = sum(row["Решение"] == "empty_name" for row in rows)

    # found = sum(row[8] in ("auto", "review") for row in rows)
    # empty = sum(row[8] == "empty_name" for row in rows)

    print(f"\nГотово: {found}/{len(rows)} найдено.")
    print(f"Пустых названий: {empty}")
    print("Результат:", args.output)


if __name__ == "__main__":
    main()
